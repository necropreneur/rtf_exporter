from openpyxl import Workbook


def write_df_to_excel(filename, df, company, serial_number, model, date, operator):

    # Always start with a new workbook
    wb = Workbook()
    ws = wb.active

    # Write individual cells in the first four rows
    ws['A1'] = company
    # Use variable for serial number
    ws['A2'] = f'Серийный номер: {serial_number}'
    ws['A3'] = f'Модель: {model}'  # Use variable for model
    ws['A4'] = 'Квитанция взвешивания от: ' + date

    # Write the header (column names) starting from row 5
    for col, column_name in enumerate(df.columns):
        ws.cell(row=5, column=col + 1, value=column_name)

    # Write the DataFrame data starting from row 6
    for index, row_data in enumerate(df.values):
        for col, data in enumerate(row_data):
            ws.cell(row=index + 6, column=col + 1, value=data)

    # Find the last row number of the df after writing it to Excel
    last_row_num = df.shape[0] + 5  # Added 5 because df starts from row 6
    # Increase the row number by 2
    desired_row_num = last_row_num + 2

    # Write "Оператор:" in cell E of the desired row
    operator_text = ''
    if operator:
        operator_text = f"Оператор: {operator}"
    ws[f'E{desired_row_num}'] = operator_text

    # Save the workbook, overwriting the existing file if it exists
    wb.save(filename)
